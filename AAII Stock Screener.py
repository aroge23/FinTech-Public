import lxml.html as web
import pandas as pd
import datetime
import pandas_datareader.data as web
import csv
from tqdm import trange as timer
import os
import openpyxl as xl
from yahoo import YahooStats
from openpyxl.utils import get_column_letter
import requests
import lxml


def GetIntrinsicValue(tick):
    tick = tick

    try:
        eps_ttm = float(YahooStats(tick)[1][11])

        url = 'https://finance.yahoo.com/quote/' + tick.upper() + '/analysis'
        html = requests.get(url)
        doc = lxml.html.fromstring(html.content)
        tables = doc.xpath('//table[@class="W(100%) M(0) BdB Bdc($seperatorColor) Mb(25px)"]')
        estimates = tables[5].xpath('.//tr[@class="BdT Bdc($seperatorColor)"]')
        growth = float(estimates[4].xpath('.//td[@class="Ta(end) Py(10px)"]')[0].xpath('./text()')[0][:-1]) / 100

        ror = 15 / 100

        mos = 50 / 100

        peratio = growth * 200
        url = 'https://www.gurufocus.com/term/pe/' + tick + '/PE-Ratio/'
        html = requests.get(url)
        doc = lxml.html.fromstring(html.content)
        tables = doc.xpath('//table[@class="R10"]')[0]
        values = tables.xpath('.//tr')[2].xpath('.//td')
        if values[-5].xpath('./text()')[0] != 'PE Ratio':
            newpe = (float(values[-1].xpath('./text()')[0]) + float(values[-5].xpath('./text()')[0])) / 2
            if peratio > newpe and newpe > 0:
                peratio = newpe
        else:
            newpe = (float(values[-1].xpath('./text()')[0]) + float(values[-4].xpath('./text()')[0])) / 2
            if peratio > newpe and newpe > 0:
                peratio = newpe

        tenyearearnings = eps_ttm * ((1 + growth) ** 9)
        fairvalue = (tenyearearnings * peratio) / ((1 + ror) ** 9)
        fairmos = fairvalue * mos
    except:
        fairvalue = 'N/A'
        pass

    return fairvalue



def FundamentalStockScreen(filepath):
    names = []
    with open(filepath, 'r', newline='') as file:
        reader = csv.reader(file)
        for row in reader:
            names.append(row[1])
    names = names[1:]

    allstocks = []
    for i in timer(len(names), desc='Fundamental Stock Screening'):
        lastyear = 0
        stock = web.DataReader(names[i], 'yahoo', lastyear, datetime.datetime.today())
        threemonth, twelvemonth = False, False
        lowclose = stock.loc[stock.index[len(stock) - 64]].Close
        for j in range(len(stock) - 63, len(stock)):
            if stock.loc[stock.index[j]].Close < lowclose:
                lowclose = stock.loc[stock.index[j]].Close
        if stock.loc[stock.index[len(stock) - 1]].Close >= (0.3 * lowclose + lowclose):
            threemonth = True

        if stock.loc[stock.index[len(stock) - 1]].Close >= 2 * stock.loc[stock.index[0]].Close:
            twelvemonth = True

        # threemonth, twelvemonth = True, True
        if threemonth and twelvemonth == True:
            allstocks.append(stock)
            values = YahooStats(names[i])[1]

            if os.path.exists('fund_stocks.csv') == False:
                with open('fund_stocks.csv', 'w', newline='') as file:
                    writer = csv.writer(file)
                    writer.writerow(['Ticker', 'Current', '52-Wk High', '52-Wk Low', '52-Wk Price Multiple'
                                     , '% Lag', 'Investment Value', '3-Month Price Gain', 'Investment Value Rank'
                                        , 'Intrinsic Value'])
                # CONVERTS CSV INTO EXCEL FOR EASY ACCESS
                read_file = pd.read_csv('fund_stocks.csv')
                read_file.to_excel('fund_stocks.xlsx', index=None, header=True)
            else:
                # CONFIGURES THE EXCEL WORKBOOK
                wb = xl.load_workbook('fund_stocks.xlsx')
                ws = wb.active

                # ADDS AN ADDITIONAL ROW OF NEW VALUES EVERY TIME YOU RUN
                current = stock.loc[stock.index[len(stock) - 1]].Close
                yearhigh, yearlow = float(values[5].split(' - ')[1]), float(values[5].split(' - ')[0])
                ws['A' + str(ws.max_row + 1)] = names[i]
                ws['B' + str(ws.max_row)] = round(current, 2)
                ws['C' + str(ws.max_row)] = yearhigh
                ws['D' + str(ws.max_row)] = yearlow
                # 52-Wk Price Multiple
                multiple = round(current / yearlow, 2)
                ws['E' + str(ws.max_row)] = multiple
                # Percent Lag
                lag = (yearhigh - (current - 0.02)) / yearhigh
                ws['F' + str(ws.max_row)] = round(lag, 2)
                # Investment Value
                inv_value = (3 * multiple) / (2 * lag)
                ws['G' + str(ws.max_row)] = round(inv_value, 2)
                # Three-Month Price Gain
                price_gain = ((current - lowclose) / lowclose) * 100
                ws['H' + str(ws.max_row)] = round(price_gain, 2)
                # Investment Value Rank
                ws['I' + str(ws.max_row)] = '=ROUND((RANK(G' + str(ws.max_row) + ', G:G' + ', 0) * 2 + 100) / H' + str(ws.max_row) + ', 2)'
                # Intrinsic Value
                ws['J' + str(ws.max_row)] = GetIntrinsicValue(names[i])


                column_widths = []
                for row in ws:
                    for i, cell in enumerate(row):
                        if len(column_widths) > i:
                            if len(str(cell.value)) > column_widths[i]:
                                column_widths[i] = len(str(cell.value)) + 5
                        else:
                            column_widths += [len(str(cell.value)) + 3]

                for i, column_width in enumerate(column_widths):
                    ws.column_dimensions[get_column_letter(i + 1)].width = column_width


                # COMMIT AND SAVE THE EXCEL WORKBOOK
                wb.save('fund_stocks.xlsx')

    # DELETE THE TEMPORARY EXCEL FILE
    # os.remove('fund_stocks.csv')


FundamentalStockScreen('/Users/anay-mac/Downloads/test.csv')

#MHH, PDD, SHOP, LOGI, GRVY, QDEL
