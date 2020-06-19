#PULLS DATA FROM YAHOO FINANCE AND WRITES TO A CSV FILE
#CREATES A NEW LINE EVERY TIME YOU RUN IT WITH DATE STAMPS
import requests
import lxml.html as web
import csv
import pandas as pd
import datetime
import openpyxl as xl
import os

def YahooStats(tick):
    url = "https://in.finance.yahoo.com/quote/" + tick.upper()

    html = requests.get(url)
    doc = web.fromstring(html.content)
    left_summary = doc.xpath('//div[@data-test="left-summary-table"]')[0]
    td = left_summary.xpath('.//td[@class="C($primaryColor) W(51%)"]')
    titles = []
    values = []
    for element in td:
        titles.append(element.xpath('.//span/text()')[0])
    td = left_summary.xpath('.//td[@class="Ta(end) Fw(600) Lh(14px)"]')
    for element in td:
        if len(element.xpath('.//span/text()')) != 0:
            values.append(element.xpath('.//span/text()')[0])
        else:
            values.append(element.xpath('./text()')[0])

    right_summary = doc.xpath('//div[@data-test="right-summary-table"]')[0]
    td = right_summary.xpath('.//td[@class="Ta(end) Fw(600) Lh(14px)"]')
    for element in td:
        if len(element.xpath('.//span/text()')) != 0:
            values.append(element.xpath('.//span/text()')[0])
        else:
            values.append(element.xpath('./text()')[0])

    td = right_summary.xpath('.//td[@class="C($primaryColor) W(51%)"]')
    for element in td:
        titles.append(element.xpath('.//span/text()')[0])
    return titles, values


# titles, values = YahooStats('aapl')
#
# #CREATES A NEW CSV FILE OF THE GIVEN STOCK IF IT DOESN'T EXIST
# tick = 'aapl'
# csv_fi = tick.upper() + '.csv'
# xlsx_fi = tick.upper() + '.xlsx'
# if os.path.exists(csv_fi) == False:
#     with open(csv_fi, 'w', newline='') as file:
#         writer = csv.writer(file)
#         writer.writerow(["Date"] + titles)
#         writer.writerow([datetime.datetime.today().strftime('%e-%b-%y')] + values)
# else:
#     #CONVERTS CSV INTO EXCEL FOR EASY ACCESS
#     read_file = pd.read_csv(csv_fi)
#     read_file.to_excel(xlsx_fi, index= None, header= True)
#
#     #CONFIGURES THE EXCEL WORKBOOK
#     wb = xl.load_workbook(xlsx_fi)
#     ws = wb.active
#     ws.title = tick.upper()
#
#     #ADDS AN ADDITIONAL ROW OF NEW VALUES EVERY TIME YOU RUN
#     col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']
#     ws[col[0] + str(ws.max_row + 1)] = datetime.datetime.today().strftime('%e-%b-%y')
#     for i in range(0, len(values)):
#         ws[col[i + 1] + str(ws.max_row)] = values[i]
#
#     #COMMIT AND SAVE THE EXCEL WORKBOOK
#     wb.save(xlsx_fi)
#
#     #CONVERTS BACK INTO CSV FILE
#     read_file = pd.read_excel(xlsx_fi)
#     read_file.to_csv(csv_fi, index= None, header= True)
#
#     #DELETE THE TEMPORARY EXCEL FILE
#     os.remove(xlsx_fi)

