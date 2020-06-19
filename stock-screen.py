import lxml.html
import openpyxl as xl
import requests
import datetime
import pandas_datareader.data as web
import os
from yahoo import YahooStats
from tqdm import trange as timer
import tqdm
from openpyxl.styles import PatternFill
import xlrd
from openpyxl.workbook import Workbook

def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)

# if os.path.exists('StockScreen.xls') == True:
cvt_xls_to_xlsx('StockScreen.xls', 'StockScreen.xlsx')
wb = xl.load_workbook('StockScreen.xlsx')
sheet = wb.active

rows = list(range(4, 24)) + list(range(26, 30)) + list(range(32, 38)) + list(range(40, 51))
for i in tqdm.tqdm(list(range(4, 24)) + list(range(26, 30)) + list(range(32, 38)) + list(range(40, 51)), desc='Processing'):
    tick = sheet['B' + str(i)].value

    stock = tick.split('.')
    if len(stock) > 1:
        stock = tick.split('.')[0] + '-' + tick.split('.')[1]
        data = web.DataReader(stock.upper(), 'yahoo', str(sheet['C3'].value), str(sheet['D3'].value))
        titles, values = YahooStats(stock)
    else:
        data = web.DataReader(tick.upper(), 'yahoo', str(sheet['C3'].value), str(sheet['D3'].value))
        titles, values = YahooStats(tick)

    sheet['C' + str(i)] = data.Close[0]
    sheet['D' + str(i)] = data.Close[len(data.Close) - 1]

    #WRITES 52-WEEK HIGH
    # sheet['C' + str(i)] = values[5][values[5].index('-') + 2:]



    #WRITES % DD
    if i >= 6:
        url = "https://in.finance.yahoo.com/quote/" + tick.upper()
        doc = lxml.html.fromstring(requests.get(url).content)
        right_summary = doc.xpath('//div[@data-test="right-summary-table"]')[0]
        row = right_summary.xpath('.//td[@class="C($primaryColor) W(51%)"]')[5]

        if 'yield' in row.xpath('.//span/text()')[0]:
            sheet['F' + str(i)] = values[13][values[13].index('(') + 1: len(values[13]) - 1]
        else:
            sheet['F' + str(i)] = 'N/A'

    #WRITES PUT/CALL
    if i >= 6:
        url = 'https://www.alphaquery.com/stock/' + str(tick.upper()) + '/volatility-option-statistics/30-day/put-call-ratio-volume'
        doc = lxml.html.fromstring(requests.get(url).content)
        below_chart = doc.xpath('//div[@id="below-chart-text"]')[0]
        sheet['I' + str(i)] = below_chart.xpath('.//strong/text()')[0]

    # WRITES PREVIOUS CLOSE
    sheet['E' + str(i)] = values[0]

    #WRITES CURRENT PE
    if i >= 6:
        sheet['K' + str(i)] = values[10]

    # Writes ROIC
    if i >= 6:
        url = 'https://www.gurufocus.com/term/ROIC/' + str(tick.upper()) + '/ROIC-Percentage/'
        doc = lxml.html.fromstring(requests.get(url).content)
        header = doc.xpath('//div[@id="def_body_detail_height"]')[0]
        sheet['L' + str(i)] = header.xpath('.//font/text()')[0].split(' ')[1]


sheet['C2'] = datetime.datetime.today().strftime('%e-%b-%y')
highlight = PatternFill(start_color='FAF60C', end_color='FAF60C', fill_type='solid')
sheet['B2'].fill, sheet['E3'].fill, sheet['I3'].fill, sheet['K3'].fill, sheet['L3'].fill = highlight, highlight, highlight, \
                                                                                           highlight, highlight



wb.save('StockScreen.xlsx')